# -*- coding: utf-8 -*-
"""
Etape 6 : export final.
  data/out/modeles.csv          - table principale (source de verite machine)
  data/out/marques.csv          - table des marques
  data/out/duels.csv            - matchs "A vs B" generes (plan editorial palier B)
  data/out/base_motos.xlsx      - classeur multi-feuilles pour travail manuel
"""
import json, os, csv, re, math
from collections import Counter, defaultdict

RAW = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "out")
os.makedirs(OUT, exist_ok=True)

rows = json.load(open(os.path.join(RAW, "normalise_qa.json"), encoding="utf-8"))

# --- signal de notoriete (etape 07) ------------------------------------
NOTO = {}
_p = os.path.join(RAW, "notoriete.json")
if os.path.exists(_p):
    NOTO = json.load(open(_p, encoding="utf-8"))
for r in rows:
    n = NOTO.get(r["titre_wikipedia"], {})
    r["vues_60j"] = n.get("vues_60j", 0)
    r["nb_langues"] = n.get("nb_langues", 0)
    r["a_version_fr"] = n.get("a_version_fr", "non")
    r["nb_langues_euro"] = n.get("nb_langues_euro", 0)

# --- enrichissement francophone (etapes 09-11) --------------------------
ENR = {}
_pe = os.path.join(RAW, "enrichissement.json")
if os.path.exists(_pe):
    ENR = json.load(open(_pe, encoding="utf-8"))

LIC_LIBRES = ("cc by", "cc0", "public domain", "attribution")
for r in rows:
    e = ENR.get(r["titre_wikipedia"], {})
    r["titre_fr"] = e.get("titre_fr", "")
    r["nom_fr"] = e.get("nom_fr", "")
    r["type_fr"] = e.get("type_fr", "")
    r["prix_lancement_eur"] = e.get("prix_lancement_eur", "")
    r["prix_source"] = e.get("prix_source", "")
    r["resume_fr"] = e.get("resume_fr", "")
    r["image_vignette"] = e.get("image_vignette", "") or r.get("image_url", "")
    r["image_licence"] = e.get("image_licence", "")
    r["image_auteur"] = e.get("image_auteur", "")
    r["image_page"] = e.get("image_page", "")
    r["url_wikipedia_fr"] = ("https://fr.wikipedia.org/wiki/"
                             + e["titre_fr"].replace(" ", "_")) if e.get("titre_fr") else ""
    lic = (r["image_licence"] or "").lower()
    if not lic:
        r["image_utilisable"] = "inconnu"
    elif any(s in lic for s in LIC_LIBRES) and "free use" not in lic:
        r["image_utilisable"] = "oui"
    else:
        r["image_utilisable"] = "verifier"

# --- specifications Wikipedia FR (etape 13) -----------------------------
# Comblement des trous uniquement : une valeur deja presente n'est jamais
# ecrasee. Les descriptions FR remplacent en revanche l'anglais, qui n'etait
# traduit qu'approximativement a l'affichage.
SPECS_FR = {}
_ps = os.path.join(RAW, "specs_fr.json")
if os.path.exists(_ps):
    SPECS_FR = json.load(open(_ps, encoding="utf-8"))

TXT_FR = [("cadre_fr", "cadre"), ("transmission_fr", "transmission"),
          ("moteur_fr", "moteur")]
_comble = Counter()
for r in rows:
    sf = SPECS_FR.get(r["titre_wikipedia"])
    if not sf:
        continue
    for k in ("cylindree_cc", "puissance_ch", "puissance_tr_min", "couple_nm",
              "couple_tr_min", "poids_sec_kg", "poids_tous_pleins_kg",
              "hauteur_selle_mm", "empattement_mm", "reservoir_l",
              "vitesse_max_kmh"):
        if sf.get(k) is not None and not r.get(k):
            r[k] = sf[k]
            _comble[k] += 1
    # puissance en kW et statut A2 : a recalculer si la puissance vient d'arriver
    if r.get("puissance_ch") and not r.get("puissance_kw"):
        r["puissance_kw"] = round(float(r["puissance_ch"]) * 0.735499, 2)
    if not r.get("poids_kg"):
        r["poids_kg"] = r.get("poids_tous_pleins_kg") or r.get("poids_sec_kg")
        if r["poids_kg"]:
            r["poids_type"] = ("tous pleins faits" if r.get("poids_tous_pleins_kg")
                               else "à sec")
    # suspensions et freins : deux champs FR fusionnes en un
    susp = " / ".join(x for x in (sf.get("suspension_av_fr"),
                                  sf.get("suspension_ar_fr")) if x)
    frein = " / ".join(x for x in (sf.get("frein_av_fr"),
                                   sf.get("frein_ar_fr")) if x)
    pneus = " / ".join(x for x in (sf.get("pneu_av_fr"), sf.get("pneu_ar_fr")) if x)
    if susp:
        r["suspension"] = susp[:250]; _comble["suspension"] += 1
    if frein:
        r["freins"] = frein[:200]; _comble["freins"] += 1
    if pneus:
        r["pneus"] = pneus[:200]; _comble["pneus"] += 1
    for src, dst in TXT_FR:
        if sf.get(src):
            r[dst] = sf[src][:300]
            _comble[dst] += 1
    r["source_specs_fr"] = sf.get("source_fr", "")

# recalcul du statut A2 apres comblement
_a2 = 0
for r in rows:
    pw, poids = r.get("puissance_kw"), r.get("poids_kg")
    if not pw:
        continue
    try:
        pw = float(pw)
    except (TypeError, ValueError):
        continue
    avant = r.get("a2_compatible")
    if pw > 35:
        r["a2_compatible"], r["a2_detail"] = "non", "puissance > 35 kW"
    elif poids:
        ratio = pw / float(poids)
        if ratio <= 0.2:
            r["a2_compatible"] = "oui"
            r["a2_detail"] = "%.1f kW, %.3f kW/kg" % (pw, ratio)
        else:
            r["a2_compatible"] = "non"
            r["a2_detail"] = "rapport %.3f kW/kg > 0.2" % ratio
    else:
        r["a2_compatible"], r["a2_detail"] = "à vérifier", "poids inconnu"
    if avant != r["a2_compatible"]:
        _a2 += 1

# recalcul de la completude
_KEY = ["cylindree_cc", "puissance_kw", "couple_nm", "poids_kg", "hauteur_selle_mm",
        "empattement_mm", "reservoir_l", "vitesse_max_kmh", "categorie",
        "annee_debut", "architecture", "transmission", "freins", "pneus", "image_url"]
for r in rows:
    r["completude_pct"] = round(
        100 * sum(1 for k in _KEY if r.get(k) not in (None, "", 0)) / len(_KEY))

if _comble:
    print("comblement FR : %s"
          % ", ".join("%s=%d" % (k, v) for k, v in _comble.most_common(8)))
    print("               %d statuts A2 recalcules" % _a2)

# marques effectivement distribuees en France (reseau officiel)
MARQUES_FR = {
 "Honda", "Yamaha", "Suzuki", "Kawasaki", "BMW", "KTM", "Ducati",
 "Triumph Motorcycles Ltd", "Triumph Engineering", "Aprilia", "Moto Guzzi",
 "Harley-Davidson", "Royal Enfield", "MV Agusta", "Husqvarna", "Benelli",
 "Piaggio", "Zero Motorcycles", "Indian Motorcycles", "Peugeot", "Kymco",
 "Norton", "BSA", "Bimota", "Buell", "Cagiva", "Derbi", "Gilera", "Laverda",
 "Vincent", "Velocette", "AJS", "Ariel", "Brough Superior", "MZ", "NSU",
 "Sachs", "Maico", "Puch", "Jawa", "Malaguti", "Mondial", "Hyosung",
}

# --- modeles collectes directement sur Wikipedia FR (etapes 14-15) -------
# Ajoutes APRES les boucles d'enrichissement : ils portent deja leurs propres
# valeurs (specs, image, audience) et ne doivent pas etre ecrases.
_sources_sup = []
for _f in ("modeles_fr_pret.json", "modeles_manuels.json"):
    _p2 = os.path.join(RAW, _f)
    if os.path.exists(_p2):
        _sources_sup += json.load(open(_p2, encoding="utf-8"))

_pf = os.path.join(RAW, "modeles_fr_pret.json")
if _sources_sup:
    _fr = _sources_sup
    _ids = {r["modele_id"] for r in rows}
    _par_id = {r["modele_id"]: r for r in rows}
    _ajout, _remplaces = [], 0
    for r in _fr:
        mid = r["modele_id"]
        if mid in _ids:
            # une fiche saisie a la main est verifiee : elle remplace une fiche
            # automatique du meme modele, jamais l'inverse
            if r.get("saisie_manuelle") and not _par_id[mid].get("saisie_manuelle"):
                _par_id[mid].update(r)
                _remplaces += 1
            continue
        _ids.add(mid)
        _ajout.append(r)
    if _remplaces:
        print("saisie manuelle : %d fiche(s) existante(s) remplacee(s)" % _remplaces)
    # aligner les cles sur le schema existant : une cle absente casserait
    # les traitements en aval, une cle vide est traitee comme "donnee absente"
    _cles = set().union(*(set(r) for r in rows)) if rows else set()
    for _r in _ajout:
        for _k in _cles:
            _r.setdefault(_k, "")
    rows.extend(_ajout)
    print("collecte FR : +%d modeles (total %d)" % (len(_ajout), len(rows)))


# normalisation des libelles : certaines fiches collectees avant la
# correction des accents portent encore "francaise"
_ACCENTS = {"francaise": "française", "suedoise": "suédoise",
            "coreenne": "coréenne", "tcheque": "tchèque",
            "americaine": "américaine"}
for r in rows:
    if r.get("ecole") in _ACCENTS:
        r["ecole"] = _ACCENTS[r["ecole"]]

def nom_affiche(r):
    """Evite 'BMW BMW G310R' : le champ 'modele' contient souvent deja la marque."""
    mo = (r.get("modele") or "").strip()
    ma = (r.get("marque") or "").strip()
    if not ma:
        return mo
    court = ma.split()[0]
    if mo.lower().startswith(ma.lower()) or mo.lower().startswith(court.lower()):
        return mo
    return "%s %s" % (court, mo)

def nom_fr_propre(r, ref):
    """Nom francais, mais jamais moins precis que l'anglais.

    Le titre d'article FR est plus fiable que le champ 'Nom' de l'infobox, qui
    tronque parfois la designation ('Yamaha YZF' au lieu de 'Yamaha YZF-R1').
    On rejette tout candidat qui perd de l'information par rapport a la
    reference anglophone.
    """
    def cle(x):
        return re.sub(r"[^a-z0-9]", "", (x or "").lower())
    kref = cle(ref)
    for cand in (r.get("titre_fr"), r.get("nom_fr")):
        c = (cand or "").strip()
        if not c or not (3 <= len(c) <= 60):
            continue
        if c.lower().startswith(("liste ", "categorie", "catégorie")):
            continue
        kc = cle(c)
        # candidat strictement moins informatif que la reference : on refuse
        if kc and kref.startswith(kc) and len(kc) < len(kref):
            continue
        # nettement plus court = designation tronquee ou obsolete
        # ('Honda GL' pour Gold Wing, 'Kawasaki ER-6' pour Ninja 650)
        if kc and kref and len(kc) < 0.7 * len(kref):
            continue
        return c
    return ""

for r in rows:
    r["nom_complet"] = nom_affiche(r)
    r["nom_affichage"] = nom_fr_propre(r, r["nom_complet"]) or r["nom_complet"]
    r["marche_fr"] = "oui" if r.get("marque") in MARQUES_FR else "non"
    # score de priorite editoriale : notoriete + pertinence FR + qualite fiche
    s = 12 * math.log1p(r["vues_60j"])
    s += 25 if r["a_version_fr"] == "oui" else 0
    s += 3 * r["nb_langues_euro"]
    s += 15 if r["marche_fr"] == "oui" else 0
    s += 0.35 * r["completude_pct"]
    a = r.get("annee_debut") or 0
    s += 18 if a >= 2015 else (9 if a >= 2005 else 0)
    r["priorite"] = round(s, 1)

# --- desambiguisation des homonymes -------------------------------------
# Certains noms couvrent deux machines distinctes (YZF-R7 1999 et 2022,
# Tiger 900 1993 et 2020...). On ajoute le millesime pour les distinguer.
_homonymes = defaultdict(list)
for r in rows:
    _homonymes[r["nom_affichage"]].append(r)
for nom, lot in _homonymes.items():
    if len(lot) < 2:
        continue
    # on ne desambiguise que si plusieurs homonymes seront reellement publies :
    # sinon la seule fiche visible porterait un millesime sans raison apparente
    publiables = [r for r in lot if (r.get("completude_pct") or 0) >= 45]
    if len(publiables) < 2:
        continue
    for r in lot:
        if r.get("annee_debut"):
            r["nom_affichage"] = "%s (%s)" % (nom, r["annee_debut"])

COLS = ["modele_id", "marque", "marque_id", "modele", "nom_complet",
        "nom_affichage", "nom_fr", "titre_fr", "type_fr",
        "prix_lancement_eur", "prix_source", "marche_fr",
        "priorite", "vues_60j", "a_version_fr", "nb_langues", "titre_wikipedia",
        "pays_origine", "ecole", "categorie", "annee_debut", "annee_fin",
        "cylindree_cc", "architecture", "refroidissement",
        "puissance_ch", "puissance_kw", "puissance_tr_min",
        "couple_nm", "couple_tr_min",
        "poids_kg", "poids_type", "poids_sec_kg", "poids_tous_pleins_kg",
        "hauteur_selle_mm", "empattement_mm", "longueur_mm", "largeur_mm", "hauteur_mm",
        "reservoir_l", "vitesse_max_kmh",
        "a2_compatible", "a2_detail",
        "moteur", "alesage_course", "compression", "transmission", "cadre",
        "suspension", "freins", "pneus", "consommation",
        "predecesseur", "successeur", "aussi_appele",
        "image_url", "wikidata_id", "url_wikipedia",
        "url_wikipedia_fr", "source_specs_fr", "image_vignette", "image_licence", "image_auteur", "image_page",
        "image_utilisable", "resume_fr",
        "production_raw", "completude_pct", "alertes", "source"]

# ------------------------------------------------------------- modeles.csv
rows.sort(key=lambda x: -x["priorite"])
with open(os.path.join(OUT, "modeles.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=COLS, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in COLS})
print("modeles.csv        : %d lignes" % len(rows))

# ------------------------------------------------------------- marques.csv
marques = defaultdict(lambda: {"n": 0, "pays": "", "ecole": "", "cat": Counter(),
                               "min": None, "max": None})
for r in rows:
    if not r["marque"]:
        continue
    m = marques[r["marque"]]
    m["n"] += 1
    m["pays"] = r["pays_origine"]
    m["ecole"] = r["ecole"]
    if r["categorie"]:
        m["cat"][r["categorie"]] += 1
    if r["annee_debut"]:
        m["min"] = r["annee_debut"] if m["min"] is None else min(m["min"], r["annee_debut"])
        m["max"] = r["annee_debut"] if m["max"] is None else max(m["max"], r["annee_debut"])

MCOLS = ["marque_id", "marque", "pays_origine", "ecole", "nb_modeles",
         "categorie_dominante", "premiere_annee", "derniere_annee"]
mrows = []
for nom, d in sorted(marques.items(), key=lambda x: -x[1]["n"]):
    mrows.append({"marque_id": re.sub(r"[^a-z0-9]+", "-", nom.lower()).strip("-"),
                  "marque": nom, "pays_origine": d["pays"], "ecole": d["ecole"],
                  "nb_modeles": d["n"],
                  "categorie_dominante": d["cat"].most_common(1)[0][0] if d["cat"] else "",
                  "premiere_annee": d["min"] or "", "derniere_annee": d["max"] or ""})
with open(os.path.join(OUT, "marques.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=MCOLS)
    w.writeheader()
    w.writerows(mrows)
print("marques.csv        : %d lignes" % len(mrows))

# ------------------------------------------------------------- duels.csv
# Regles : meme categorie, cylindree proche (+/-25%), meme statut A2,
# les deux fiches suffisamment completes. Bonus si ecoles differentes.
elig = [r for r in rows
        if r["categorie"] and r["cylindree_cc"] and r["completude_pct"] >= 50
        and r["marque"] and (r["annee_debut"] or 0) >= 1990
        and r["marche_fr"] == "oui"
        and r["vues_60j"] >= 150
        and r["categorie"] not in ("Competition", "Minibike", "Trois-roues")]
print("duels : %d modeles eligibles" % len(elig))

duels = []
seen = set()
for i, a in enumerate(elig):
    for b in elig[i+1:]:
        if a["categorie"] != b["categorie"]:
            continue
        if a["marque"] == b["marque"]:
            continue
        ca, cb = a["cylindree_cc"], b["cylindree_cc"]
        if min(ca, cb) / max(ca, cb) < 0.75:
            continue
        if a["a2_compatible"] != b["a2_compatible"]:
            continue
        key = tuple(sorted([a["modele_id"], b["modele_id"]]))
        if key in seen:
            continue
        seen.add(key)
        # un duel ne vaut que par son cote le plus faible -> min() sur la notoriete
        score = min(a["priorite"], b["priorite"])
        score += 0.15 * (a["completude_pct"] + b["completude_pct"]) / 2
        if a["ecole"] != b["ecole"] and a["ecole"] and b["ecole"]:
            score += 20          # duel inter-ecoles = angle editorial
        if a["a_version_fr"] == "oui" and b["a_version_fr"] == "oui":
            score += 15          # les deux connus du public francophone
        # --- pertinence marche actuel -----------------------------------
        # Sans ce reequilibrage, les sportives des annees 1990-2000 dominent :
        # leurs articles Wikipedia sont tres complets, mais l'intention d'achat
        # francaise est ailleurs.
        rec = max(a["annee_debut"] or 0, b["annee_debut"] or 0)
        if rec >= 2018:
            score += 40
        elif rec >= 2012:
            score += 25
        elif rec >= 2005:
            score += 8
        else:
            score -= 15
        # encore au catalogue = intention d'achat vivante
        encore = sum(1 for x in (a, b) if not x["annee_fin"])
        score += 18 * encore
        # arretee depuis longtemps : interet historique, pas commercial
        for x in (a, b):
            if x["annee_fin"]:
                try:
                    if int(x["annee_fin"]) < 2010:
                        score -= 20
                except ValueError:
                    pass
        if a["a2_compatible"] == "oui":
            score += 18          # trafic permis A2 = fort volume FR
        na, nb = a["nom_affichage"], b["nom_affichage"]
        duels.append({
            "duel_id": "%s-vs-%s" % key,
            "modele_a_id": a["modele_id"], "modele_a": na,
            "modele_b_id": b["modele_id"], "modele_b": nb,
            "titre_page": "%s ou %s : lequel choisir ?" % (na, nb),
            "categorie": a["categorie"],
            "ecole_a": a["ecole"], "ecole_b": b["ecole"],
            "inter_ecoles": "oui" if (a["ecole"] != b["ecole"] and a["ecole"] and b["ecole"]) else "non",
            "cylindree_a": ca, "cylindree_b": cb,
            "puissance_ch_a": a["puissance_ch"] or "", "puissance_ch_b": b["puissance_ch"] or "",
            "poids_a": a["poids_kg"] or "", "poids_b": b["poids_kg"] or "",
            "a2": a["a2_compatible"],
            "score_priorite": round(score, 1),
        })

duels.sort(key=lambda d: -d["score_priorite"])

# --- plafond par modele -------------------------------------------------
# Sans plafond, une poignee de modeles monopolise la page (jusqu'a 18 duels
# pour la YZF-R7). On parcourt du meilleur au moins bon et on retient tant
# qu'aucun des deux modeles n'a atteint son quota.
MAX_PAR_MODELE = 6
compte = Counter()
retenus = []
for d in duels:
    a_id, b_id = d["modele_a_id"], d["modele_b_id"]
    if compte[a_id] >= MAX_PAR_MODELE or compte[b_id] >= MAX_PAR_MODELE:
        continue
    compte[a_id] += 1
    compte[b_id] += 1
    retenus.append(d)
print("duels : %d generes -> %d apres plafond de %d par modele"
      % (len(duels), len(retenus), MAX_PAR_MODELE))
duels = retenus
DCOLS = list(duels[0].keys()) if duels else []
with open(os.path.join(OUT, "duels.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=DCOLS)
    w.writeheader()
    w.writerows(duels)
print("duels.csv          : %d lignes (top 4000)" % len(duels))

# ------------------------------------------------------------- XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
H_FILL = PatternFill("solid", fgColor="123B41")
H_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="123B41")
THIN = Side(style="thin", color="D0D6D5")

def style_sheet(ws, ncols, widths=None, freeze="A2"):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for i in range(1, ncols + 1):
        L = get_column_letter(i)
        ws.column_dimensions[L].width = (widths or {}).get(i, 16)

# --- feuille LISEZ-MOI
ws = wb.active
ws.title = "LISEZ-MOI"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95
LISEZ = [
 ("BASE DE DONNEES MOTOS", ""),
 ("", ""),
 ("Genere le", "24 aout 2026"),
 ("Modeles", str(len(rows))),
 ("Marques", str(len(mrows))),
 ("Duels generes", str(len(duels))),
 ("Completude moyenne", "%.0f%%" % (sum(r["completude_pct"] for r in rows) / len(rows))),
 ("", ""),
 ("SOURCE", ""),
 ("Origine des donnees", "Wikipedia anglophone - infobox 'Motorcycle', extraites via l'API MediaWiki"),
 ("Licence", "CC BY-SA 4.0 - la reutilisation impose d'CREDITER Wikipedia et de partager"),
 ("", "les contenus derives sous la meme licence. Un lien vers l'article source est"),
 ("", "fourni dans la colonne url_wikipedia pour chaque ligne : conserve-le."),
 ("Point juridique", "Les donnees factuelles brutes (cylindree, poids) ne sont pas protegeables"),
 ("", "en elles-memes. La structure de la base l'est. Ne recopie jamais un texte"),
 ("", "de description : reecris-le."),
 ("", ""),
 ("FIABILITE", ""),
 ("Controle effectue", "9 modeles de reference verifies contre les fiches constructeur :"),
 ("", "9 valeurs conformes, 0 ecart superieur a 6%."),
 ("Nettoyage", "141 valeurs aberrantes detectees et EFFACEES (jamais corrigees a l'aveugle)."),
 ("Regle appliquee", "Aucune valeur n'a ete inventee ou estimee. Une cellule vide signifie"),
 ("", "que la donnee est absente de la source, pas qu'elle vaut zero."),
 ("Colonne 'alertes'", "Signale les incoherences residuelles a verifier a la main."),
 ("Colonne 'completude_pct'", "Part des 15 champs cles remplis. Trie dessus pour prioriser ton travail."),
 ("", ""),
 ("LIMITES CONNUES", ""),
 ("Articles de famille", "Certains articles couvrent plusieurs generations (ex. Ducati Monster) :"),
 ("", "l'infobox utilise alors un tableau, et les champs chiffres ressortent vides."),
 ("", "Ces fiches sont a completer manuellement, generation par generation."),
 ("Millesimes", "La base est au niveau MODELE, pas au niveau ANNEE-MODELE. Pour un"),
 ("", "comparateur fin, il faudra eclater les generations."),
 ("Prix", "Absent : Wikipedia ne le documente pas. A saisir manuellement ou via"),
 ("", "les tarifs constructeurs."),
 ("Marche francais", "Base anglophone : certains modeles vendus en France manquent, et les"),
 ("", "noms sont les designations techniques (Africa Twin = CRF1000L)."),
 ("", ""),
 ("ORDRE DE TRAVAIL CONSEILLE", ""),
 ("1", "Filtrer completude_pct >= 80 -> 398 fiches quasi pretes."),
 ("2", "Parmi elles, garder les modeles vendus en France depuis 2015 -> ton lot 1."),
 ("3", "Verifier la colonne 'alertes' sur ce lot, puis publier."),
 ("4", "Ouvrir la feuille 'Duels' et prendre les 20 premieres lignes comme plan editorial."),
]
for i, (a, b) in enumerate(LISEZ, 1):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical="top")
    if a and not b:
        ws.cell(row=i, column=1).font = TITLE_FONT
    elif a:
        ws.cell(row=i, column=1).font = Font(bold=True, size=10)
ws["A1"].font = Font(bold=True, size=18, color="123B41")

# --- feuille Modeles
ws = wb.create_sheet("Modeles")
ws.append(COLS)
for r in rows:
    ws.append([("" if r.get(k) is None else r.get(k)) for k in COLS])
W = {COLS.index("modele") + 1: 30, COLS.index("titre_wikipedia") + 1: 30,
     COLS.index("moteur") + 1: 45, COLS.index("suspension") + 1: 32,
     COLS.index("freins") + 1: 32, COLS.index("pneus") + 1: 28,
     COLS.index("transmission") + 1: 28, COLS.index("cadre") + 1: 26,
     COLS.index("image_url") + 1: 30, COLS.index("url_wikipedia") + 1: 34,
     COLS.index("a2_detail") + 1: 22, COLS.index("alertes") + 1: 34,
     COLS.index("production_raw") + 1: 18, COLS.index("categorie") + 1: 20,
     COLS.index("architecture") + 1: 20, COLS.index("marque") + 1: 18}
style_sheet(ws, len(COLS), W)
# coloration completude
ci = COLS.index("completude_pct") + 1
for row in range(2, ws.max_row + 1):
    v = ws.cell(row=row, column=ci).value
    if isinstance(v, (int, float)):
        if v >= 80:
            col = "C6E7D0"
        elif v >= 60:
            col = "FDF0CE"
        else:
            col = "F7D6D0"
        ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=col)

# --- feuille Marques
ws = wb.create_sheet("Marques")
ws.append(MCOLS)
for r in mrows:
    ws.append([r[k] for k in MCOLS])
style_sheet(ws, len(MCOLS), {1: 22, 2: 26, 3: 16, 4: 16, 5: 12, 6: 22, 7: 14, 8: 14})

# --- feuille Duels
ws = wb.create_sheet("Duels")
if duels:
    ws.append(DCOLS)
    for d in duels[:1500]:
        ws.append([d[k] for k in DCOLS])
    style_sheet(ws, len(DCOLS),
                {DCOLS.index("modele_a") + 1: 30, DCOLS.index("modele_b") + 1: 30,
                 DCOLS.index("titre_page") + 1: 55, DCOLS.index("duel_id") + 1: 34,
                 DCOLS.index("modele_a_id") + 1: 26, DCOLS.index("modele_b_id") + 1: 26,
                 DCOLS.index("categorie") + 1: 20})

# --- feuille Dictionnaire
ws = wb.create_sheet("Dictionnaire")
DICO = [
 ("Colonne", "Type", "Unite", "Description"),
 ("modele_id", "texte", "-", "Identifiant unique, utilisable comme slug d'URL"),
 ("marque / marque_id", "texte", "-", "Constructeur et son slug"),
 ("modele", "texte", "-", "Nom commercial du modele"),
 ("nom_complet", "texte", "-", "Marque + modele sans doublon (source anglophone)"),
 ("nom_affichage", "texte", "-", "NOM A AFFICHER : francais si dispo, sinon anglais"),
 ("nom_fr", "texte", "-", "Nom commercial francais (infobox Wikipedia FR)"),
 ("type_fr", "texte", "-", "Type/categorie en francais (Wikipedia FR)"),
 ("prix_lancement_eur", "entier", "EUR", "Prix au lancement - devise verifiee. 46 modeles seulement"),
 ("prix_source", "texte", "-", "Provenance et annee du prix"),
 ("resume_fr", "texte", "-", "Intro Wikipedia FR - AMORCE, a REECRIRE (CC BY-SA)"),
 ("image_vignette", "URL", "-", "Vignette 800px - A UTILISER sur le site (l'original pese jusqu'a 1 Mo)"),
 ("image_licence", "texte", "-", "Licence de l'image (Commons)"),
 ("image_auteur", "texte", "-", "Auteur a crediter"),
 ("image_utilisable", "texte", "-", "oui / verifier / inconnu - controle avant publication"),
 ("marche_fr", "texte", "-", "oui = marque distribuee en France. FILTRE PRINCIPAL pour ton lot 1"),
 ("pays_origine", "texte", "-", "Pays du constructeur"),
 ("ecole", "texte", "-", "japonaise / italienne / americaine / britannique / allemande..."),
 ("categorie", "texte", "-", "Roadster, Sportive, Trail, Custom, Scooter, Routiere..."),
 ("annee_debut / annee_fin", "entier", "annee", "Periode de production. Fin vide = encore produit"),
 ("cylindree_cc", "decimal", "cm3", "Cylindree totale"),
 ("architecture", "texte", "-", "Monocylindre, Twin parallele, V-twin, 4 en ligne..."),
 ("refroidissement", "texte", "-", "Air, Liquide, Air/huile"),
 ("puissance_ch", "decimal", "ch (PS)", "Puissance en chevaux metriques - usage francais"),
 ("puissance_kw", "decimal", "kW", "Puissance en kilowatts - sert au calcul A2"),
 ("puissance_tr_min", "entier", "tr/min", "Regime de puissance maxi"),
 ("couple_nm", "decimal", "Nm", "Couple maxi"),
 ("poids_kg", "decimal", "kg", "Poids retenu : tous pleins faits si dispo, sinon a sec"),
 ("poids_type", "texte", "-", "Precise lequel des deux a ete retenu"),
 ("hauteur_selle_mm", "decimal", "mm", "Hauteur de selle - critere d'achat majeur"),
 ("empattement_mm", "decimal", "mm", "Distance entre axes de roues"),
 ("reservoir_l", "decimal", "L", "Capacite du reservoir"),
 ("vitesse_max_kmh", "decimal", "km/h", "Vitesse maximale annoncee"),
 ("a2_compatible", "texte", "-", "oui / non / a verifier - calcule : <=35 kW ET <=0,2 kW/kg"),
 ("a2_detail", "texte", "-", "Justification du calcul A2"),
 ("image_url", "URL", "-", "Image principale de l'article Wikipedia"),
 ("url_wikipedia", "URL", "-", "Article source - A CONSERVER pour l'attribution CC BY-SA"),
 ("completude_pct", "entier", "%", "Part des 15 champs cles remplis"),
 ("priorite", "decimal", "-", "Score editorial : notoriete + pertinence FR + qualite. TRIER DESSUS"),
 ("vues_60j", "entier", "vues", "Consultations de l'article Wikipedia EN sur 60 jours"),
 ("a_version_fr", "texte", "-", "Une page Wikipedia francaise existe : proxy de notoriete en France"),
 ("nb_langues", "entier", "-", "Nombre de versions linguistiques de l'article"),
 ("alertes", "texte", "-", "Incoherences detectees, a verifier manuellement"),
]
for r in DICO:
    ws.append(list(r))
style_sheet(ws, 4, {1: 26, 2: 12, 3: 12, 4: 72})

# --- feuille Controle qualite
ws = wb.create_sheet("A verifier")
ws.append(["modele_id", "marque", "modele", "completude_pct", "alertes", "url_wikipedia"])
n_al = 0
for r in rows:
    if r.get("alertes"):
        ws.append([r["modele_id"], r["marque"], r["modele"], r["completude_pct"],
                   r["alertes"], r["url_wikipedia"]])
        n_al += 1
style_sheet(ws, 6, {1: 30, 2: 18, 3: 30, 4: 14, 5: 60, 6: 40})

wb.save(os.path.join(OUT, "base_motos.xlsx"))
print("base_motos.xlsx    : 6 feuilles (%d lignes a verifier)" % n_al)

print("\n=== TOP 15 DUELS PRIORITAIRES ===")
for d in duels[:15]:
    print("  %5.1f  %-58s %s" % (d["score_priorite"], d["titre_page"][:58],
                                 "[inter-ecoles]" if d["inter_ecoles"] == "oui" else ""))
